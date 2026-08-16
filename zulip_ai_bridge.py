#!/usr/bin/env python3
"""
Zulip to local AI CLI bridge.
Listens for messages on Zulip, runs the local `ai` agent, and posts responses back.

Each message is dispatched to its own daemon thread so the Zulip listener loop
is never blocked — you can send new messages while a previous one is still being
processed by the agent.
"""

import subprocess
import threading
import sys
import os
import logging
try:
    import zulip
except ImportError:
    import types
    zulip = types.ModuleType("zulip")
    zulip.Client = None
    sys.modules["zulip"] = zulip
import re
import requests
from urllib.parse import unquote, quote
import csv
import io
import time

# Configure logging for truncation events
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("ZulipBridge")

# Maximum file size to process (10 MB)
MAX_FILE_SIZE = 10 * 1024 * 1024

# MIME type to content handler mapping
TEXT_EXTS = {'.txt', '.md', '.rst', '.py', '.js', '.ts', '.jsx', '.tsx',
             '.c', '.h', '.cpp', '.hpp', '.java', '.go', '.rs', '.rb',
             '.sh', '.bash', '.zsh', '.bat', '.cmd', '.ps1',
             '.yaml', '.yml', '.toml', '.ini', '.cfg', '.conf',
             '.json', '.xml', '.html', '.htm', '.css', '.scss',
             '.sql', '.r', '.m', '.mm', '.swift', '.kt', '.kts'}

IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp', '.tiff'}

PDF_EXTS = {'.pdf'}

SPREADSHEET_EXTS = {'.xlsx', '.xls', '.csv'}

ARCHIVE_EXTS = {'.zip', '.tar', '.gz', '.bz2', '.xz', '.7z'}


class FileParser:
    """
    Downloads Zulip uploads and extracts text content for various file types.
    """

    def __init__(self, client, base_url):
        self.client = client
        self.base_url = base_url.rstrip('/')
        # Try to import optional dependencies
        self._pdfplumber = None
        self._tesseract = None
        self._pillow = None
        try:
            import pdfplumber
            self._pdfplumber = pdfplumber
        except ImportError:
            pass
        try:
            import pypdfium2 as pdfium
            self._pdfium = pdfium
        except ImportError:
            pass
        try:
            from PIL import Image
            self._pillow = Image
        except ImportError:
            pass
        try:
            import pytesseract
            self._tesseract = pytesseract
        except ImportError:
            pass

    def _download_file(self, url, dest_path):
        """Download a file with retry and size checking."""
        # Validate URL is from a trusted domain (the configured Zulip server)
        if not self._is_trusted_url(url):
            logger.warning(f"Blocked download from untrusted domain: {url}")
            return False, "Download from untrusted domain"

        try:
            resp = requests.get(url, stream=True, timeout=30)
            # The final URL after following redirects must also be trusted.
            if not self._is_trusted_redirect(resp.url, url):
                logger.warning(f"Redirect to untrusted domain: {resp.url}")
                return False, "Download redirects to untrusted domain"
            resp.raise_for_status()

            # Check Content-Length header if available
            content_length = int(resp.headers.get('Content-Length', 0))
            if content_length > MAX_FILE_SIZE:
                logger.warning(f"File too large: {content_length} bytes (max {MAX_FILE_SIZE})")
                return False, "File exceeds size limit"

            with open(dest_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    f.write(chunk)
            return True, None
        except requests.exceptions.RequestException as e:
            logger.error(f"Download failed: {e}")
            return False, str(e)
        except TimeoutError as e:
            logger.error(f"Download timed out: {e}")
            return False, str(e)
        except OSError as e:
            logger.error(f"Download file system error: {e}")
            return False, str(e)

    def _is_trusted_url(self, url):
        """Check if a URL is from the configured Zulip server domain.

        Accepts exact domain matches, subdomains of the allowed domain,
        and relative paths (which are always same-origin).
        """
        try:
            from urllib.parse import urlparse
            parsed = urlparse(url)
            # Relative paths (empty netloc) are same-origin and always trusted.
            if not parsed.netloc:
                return True
            # Use the stored base_url (already normalized by the constructor).
            allowed = self.base_url.replace('https://', '').replace('http://', '').rstrip('/')
            downloaded = parsed.netloc.replace('https://', '').replace('http://', '').rstrip('/')
            return (
                downloaded == allowed
                or downloaded.endswith("." + allowed)
                or allowed.endswith("." + downloaded)
            )
        except Exception:
            return False

    def _is_trusted_redirect(self, final_url, original_url):
        """Check if a redirect target is trusted.

        After following redirects, the final URL must also be from a
        trusted domain to prevent redirects to malicious sites.
        """
        try:
            from urllib.parse import urlparse
            final_parsed = urlparse(final_url)
            # If no netloc, treat as same-origin relative redirect.
            if not final_parsed.netloc:
                return True
            return self._is_trusted_url(final_url)
        except Exception:
            return False

    def _detect_file_type(self, path, content_type=None):
        """Detect file type from extension and/or MIME type."""
        ext = os.path.splitext(path)[1].lower()
        if ext in TEXT_EXTS:
            return 'text'
        elif ext in IMAGE_EXTS:
            return 'image'
        elif ext in PDF_EXTS:
            return 'pdf'
        elif ext in SPREADSHEET_EXTS:
            return 'spreadsheet'
        elif ext in ARCHIVE_EXTS:
            return 'archive'
        # Fallback: try to detect from content
        try:
            with open(path, 'rb') as f:
                header = f.read(512)
            if header.startswith(b'PK'):  # ZIP/DOCX/XLSX/PPTX
                return 'archive' if ext == '' else 'binary'
            elif header.startswith(b'%PDF'):
                return 'pdf'
            elif header.startswith(b'{') or header.startswith(b'['):
                return 'text'
            elif header.startswith(b'<?xml') or header.startswith(b'<'):
                return 'text'
        except Exception:
            pass
        if content_type:
            if 'json' in content_type:
                return 'text'
            elif 'xml' in content_type:
                return 'text'
            elif 'html' in content_type:
                return 'text'
        return 'unknown'

    def _parse_text(self, path):
        """Read text files."""
        try:
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read(MAX_FILE_SIZE)
            return content
        except Exception as e:
            logger.error(f"Failed to read text file {path}: {e}")
            return None

    def _parse_pdf(self, path):
        """Extract text from PDF files."""
        if self._pdfplumber:
            try:
                text = []
                with self._pdfplumber.open(path) as pdf:
                    for i, page in enumerate(pdf.pages):
                        page_text = page.extract_text()
                        if page_text:
                            text.append(f"--- Page {i+1} ---\n{page_text}")
                return '\n\n'.join(text) if text else None
            except Exception as e:
                logger.error(f"PDF extraction failed with pdfplumber: {e}")

        if self._pdfium:
            try:
                doc = self._pdfium.Document(path)
                text = []
                for i, page in enumerate(doc):
                    page_text = page.get_textpage().get_text_bounded()
                    if page_text:
                        text.append(f"--- Page {i+1} ---\n{page_text}")
                return '\n\n'.join(text) if text else None
            except Exception as e:
                logger.error(f"PDF extraction failed with pypdfium2: {e}")
                return None

        # Fallback to pypdf
        try:
            from pypdf import PdfReader
            reader = PdfReader(path)
            text = []
            for i, page in enumerate(reader.pages):
                page_text = page.extract_text()
                if page_text:
                    text.append(f"--- Page {i+1} ---\n{page_text}")
            return '\n\n'.join(text) if text else None
        except Exception as e:
            logger.error(f"PDF extraction failed: {e}")
            return None

    def _parse_image(self, path):
        """Extract text from images using OCR."""
        if self._tesseract:
            try:
                return self._tesseract.image_to_string(self._pillow.open(path))
            except Exception as e:
                logger.error(f"OCR failed: {e}")
        return None

    def _parse_csv(self, path):
        """Parse CSV files into readable format."""
        try:
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read(MAX_FILE_SIZE)
            # Parse and reformat to show structure clearly
            reader = csv.DictReader(io.StringIO(content))
            if not reader.fieldnames:
                return content
            lines = []
            lines.append(','.join(reader.fieldnames))
            for i, row in enumerate(reader):
                line = ','.join(str(row.get(f, '')) for f in reader.fieldnames)
                lines.append(line)
                if i >= 100:  # Limit to 100 rows
                    lines.append('... [truncated]')
                    break
            return '\n'.join(lines)
        except Exception as e:
            logger.error(f"CSV parsing failed: {e}")
            return content if 'content' in dir() else None

    def _parse_excel(self, path):
        """Parse Excel files."""
        if self._pillow is None:
            return None
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames[:3]:  # Limit to 3 sheets
                ws = wb[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 100:
                        text_parts.append('... [truncated]')
                        break
                    text_parts.append('\t'.join(str(c) if c is not None else '' for c in row))
            return '\n'.join(text_parts)
        except Exception as e:
            logger.error(f"Excel parsing failed: {e}")
            return None

    def _parse_archive(self, path):
        """Extract contents from archives (list files)."""
        import zipfile
        import tarfile
        import gzip as gz
        text_parts = []
        try:
            if path.endswith('.zip'):
                with zipfile.ZipFile(path, 'r') as z:
                    text_parts.extend(z.namelist())
            elif path.endswith(('.tar', '.gz', '.bz2', '.xz')):
                with tarfile.open(path, 'r:*') as t:
                    text_parts.extend(t.getnames())
            elif path.endswith('.gz'):
                try:
                    with gz.open(path, 'rt', encoding='utf-8', errors='replace') as f:
                        text_parts.append(f.read(MAX_FILE_SIZE))
                except Exception:
                    pass
            return '\n'.join(text_parts) if text_parts else None
        except Exception as e:
            logger.error(f"Archive extraction failed: {e}")
            return None

    def parse_file(self, path):
        """Parse a file and return extracted text content."""
        file_type = self._detect_file_type(path)
        logger.info(f"Parsing {os.path.basename(path)} as {file_type}")

        handlers = {
            'text': self._parse_text,
            'pdf': self._parse_pdf,
            'image': self._parse_image,
            'spreadsheet': self._parse_csv if path.endswith('.csv') else self._parse_excel,
            'archive': self._parse_archive,
        }

        handler = handlers.get(file_type)
        if handler:
            content = handler(path)
            if content:
                return content
            return f"*[Unable to extract content from {file_type} file]*"
        elif file_type == 'binary':
            return f"*[Binary file: {os.path.basename(path)} - cannot extract text]*"
        else:
            return f"*[Unsupported file type for: {os.path.basename(path)}]*"

    def process_message_urls(self, content):
        """
        Process a message content string: find Zulip upload URLs,
        download the files, extract content, and replace URLs with content.
        """
        # Find all Zulip upload URLs
        url_pattern = re.compile(r'https?://[^\s]+/user_uploads/[^\s\)]+')
        urls = url_pattern.findall(content)

        if not urls:
            return content, []

        processed_urls = []
        download_dir = os.path.join(os.path.expanduser('~'), '.cache', 'zulip_ai_uploads')
        os.makedirs(download_dir, exist_ok=True)

        for url in urls:
            # Extract filename from URL
            decoded_url = unquote(url)
            filename = os.path.basename(decoded_url.split('?')[0])
            if not filename:
                filename = 'downloaded_file'

            dest_path = os.path.join(download_dir, filename)

            # Download the file
            success, error = self._download_file(decoded_url, dest_path)
            if not success:
                content = content.replace(url, f"*⚠️ Failed to download: {error}*")
                continue

            # Parse the file content
            extracted = self.parse_file(dest_path)

            # Replace URL with extracted content or placeholder
            if extracted and not extracted.startswith('*['):
                # Format the content nicely
                content = content.replace(url, f"```[File: {filename}]\n{extracted}\n```")
            else:
                # File couldn't be parsed
                content = content.replace(url, f"*⚠️ Cannot extract text from {filename}*\n*Note: File saved at {dest_path}*")

            processed_urls.append(filename)
            logger.info(f"Processed: {filename}")

        return content, processed_urls


def load_env_file():
    """Load environment variables from ~/.local/share/ai/env."""
    env_file = os.path.expanduser("~/.local/share/ai/env")
    loaded = {}
    if os.path.exists(env_file):
        try:
            with open(env_file, "r") as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("export "):
                        line = line[7:].strip()
                    if "=" in line and not line.startswith("#"):
                        k, v = line.split("=", 1)
                        k = k.strip()
                        v = v.strip().strip("'\"")
                        loaded[k] = v
        except Exception as e:
            logger.warning(f"Could not load env file {env_file}: {e}")
    return loaded


def normalize_text(text):
    """Normalize soft-wrapped prose paragraphs while strictly preserving code blocks."""
    if not text:
        return ""

    # Split text into code blocks and non-code blocks
    parts = re.split(r'(```[\s\S]*?```)', text)
    result = []

    for part in parts:
        if part.startswith('```') and part.endswith('```'):
            # Code block - preserve exactly as is
            result.append(part)
        else:
            # Prose blocks
            paragraphs = re.split(r'\n{2,}', part)
            norm_paras = []
            for p in paragraphs:
                p_clean = p.strip()
                if not p_clean:
                    continue
                # If block is code with 4-space or tab indentation, preserve
                lines = p_clean.split('\n')
                if any(line.startswith('    ') or line.startswith('\t') for line in lines):
                    norm_paras.append(p_clean)
                else:
                    # Join lines with spaces and collapse multiple spaces
                    joined = ' '.join(line.strip() for line in lines if line.strip())
                    joined = re.sub(r' {2,}', ' ', joined)
                    if joined:
                        norm_paras.append(joined)
            result.append('\n\n'.join(norm_paras))

    return ''.join(result)


def clean_response(text):
    if not text:
        return ""
    # Strip all ANSI escape codes (colour, dim, bold, cursor movement, etc.)
    ansi_escape = re.compile(r'\x1b(?:\[[0-9;]*[a-zA-Z]|\]\d*;[^\x07]*\x07|[@-Z\\-_])')
    clean = ansi_escape.sub('', text)
    # Remove carriage returns
    clean = clean.replace('\r', '')
    # Replace long unicode horizontal lines with standard markdown horizontal rules
    clean = clean.replace("────────────────────────────────────────────", "---")
    # Collapse runs of blank lines to at most two
    clean = re.sub(r'\n{3,}', '\n\n', clean)
    # Join wrapped lines in prose while preserving code & tables
    clean = normalize_text(clean)
    return clean.strip()


class ZulipAiBridge:
    def __init__(self, client=None):
        if zulip is None and client is None:
            raise RuntimeError("The 'zulip' Python package is required. Install it via `pip install zulip`.")
        # Allow dependency injection for testing / alternative backends
        self.client = client if client is not None else zulip.Client()
        self.bot_email = self.client.email
        print(f"Loaded credentials for: {self.bot_email} on {self.client.base_url}")
        self.detected_owner = self._detect_owner()
        # Initialize the file parser for processing uploaded documents
        self._file_parser = FileParser(self.client, self.client.base_url)
        print("File parser initialized — will extract content from uploaded documents.")

    def _is_trusted_url(self, url):
        """Check if a URL is from the trusted Zulip domain."""
        return self._file_parser._is_trusted_url(url)

    def _is_trusted_redirect(self, final_url, original_url):
        """Check if a redirect final URL is trusted."""
        return self._file_parser._is_trusted_redirect(final_url, original_url)

    def _download_file(self, url, dest_path):
        """Download a file, delegating to the file parser."""
        return self._file_parser._download_file(url, dest_path)

    def _detect_owner(self):
        """Try to detect the owner's Zulip email/username from past private messages."""
        try:
            payload = {
                "anchor": "newest",
                "num_before": 20,
                "num_after": 0,
                "narrow": [{"operator": "is", "operand": "private"}],
                "apply_markdown": False
            }
            res = self.client.get_messages(payload)
            if res.get("result") == "success":
                messages = res.get("messages", [])
                for msg in reversed(messages):
                    sender_email = msg.get("sender_email")
                    if sender_email and sender_email != self.bot_email:
                        print(f"Detected owner from past private messages: {sender_email}")
                        return sender_email
        except Exception as e:
            print(f"Error detecting owner from messages: {e}")
        return None

    def _send_reply(self, msg, content):
        """Send a reply to the same stream/topic or private thread.

        Retries on transient errors (rate limiting / connection drops) with
        exponential backoff, since Zulip may temporarily refuse messages.
        """
        max_retries = 3
        backoff = 2
        message = {
            "type": "private" if msg['type'] == 'private' else "stream",
        }
        if msg['type'] == 'private':
            message["to"] = [msg['sender_email']]
        else:
            message["to"] = msg['display_recipient']
            message["topic"] = msg['subject']
        message["content"] = content

        last_error = None
        for attempt in range(max_retries):
            try:
                self.client.send_message(message)
                return
            except Exception as e:
                last_error = e
                if attempt < max_retries - 1:
                    print(f"⚠️ send_message failed (attempt {attempt+1}/{max_retries}): {e}")
                    time.sleep(backoff)
                    backoff = min(backoff * 2, 30)
                else:
                    print(f"❌ send_message failed after {max_retries} attempts: {last_error}")

        # Append an error notice to the original message thread as a last resort.
        error_notice = (
            f"⚠️ Could not deliver reply after {max_retries} attempts: {last_error}"
        )
        if msg['type'] == 'private':
            message["content"] = error_notice
        else:
            message["content"] = error_notice
        try:
            self.client.send_message(message)
        except Exception:
            pass

    def _get_context_messages(self, msg, limit=5):
        """Fetch up to `limit` previous messages in the same thread/conversation."""
        try:
            if msg['type'] == 'stream':
                narrow = [
                    {"operator": "stream", "operand": msg['display_recipient']},
                    {"operator": "topic", "operand": msg['subject']}
                ]
            else:  # private
                if isinstance(msg['display_recipient'], list):
                    emails = [r['email'] for r in msg['display_recipient']]
                    narrow = [{"operator": "pm-with", "operand": ",".join(emails)}]
                else:
                    narrow = [{"operator": "pm-with", "operand": msg['sender_email']}]

            payload = {
                "anchor": msg['id'],
                "num_before": limit,
                "num_after": 0,
                "narrow": narrow,
                "apply_markdown": False
            }
            res = self.client.get_messages(payload)
            if res.get("result") == "success":
                messages = res.get("messages", [])
                context_messages = [m for m in messages if m['id'] != msg['id']]
                return context_messages
            else:
                print(f"Zulip API error fetching context: {res.get('msg')}")
        except Exception as e:
            print(f"Error fetching context messages: {e}")
        return []

    def _construct_prompt_with_context(self, msg, content, context_messages):
        """Build a prompt that includes the context messages organically."""
        if not context_messages:
            return content

        context_window = ContextWindowManager()
        context_text = context_window.format_context(context_messages, self.bot_email)
        prompt = f"{context_text}\n\n{content}"

        return prompt

    def _ai_mode(self):
        """Resolve the AI permission mode for subprocess `ai` invocations.

        Defaults to "auto" so the bridge actually executes work instead of
        halting to ask for confirmation on every state-changing action, which
        is useless over Zulip (no interactive approve prompt). Override per
        deployment by setting the BRIDGE_AI_MODE env var (auto/plan/manual).
        """
        return os.environ.get("BRIDGE_AI_MODE", "auto").strip().lower()

    def _resolve_ai_bin(self):
        """Locate the ai binary on the host system."""
        if "INFER_BIN_PATH" in os.environ and os.path.isfile(os.environ["INFER_BIN_PATH"]):
            return os.environ["INFER_BIN_PATH"]
        repo_bin = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai")
        if os.path.isfile(repo_bin) and os.access(repo_bin, os.X_OK):
            return repo_bin
        local_bin = os.path.expanduser("~/.local/bin/ai")
        if os.path.isfile(local_bin) and os.access(local_bin, os.X_OK):
            return local_bin
        return "ai"

    def _truncate_reply(self, text, max_chars=9000):
        """Truncate a reply to fit Zulip's per-message size limit."""
        if len(text) <= max_chars:
            return text
        head = text[:max_chars].rstrip()
        return f"{head}\n\n*…[reply truncated — {len(text) - max_chars} chars omitted]*"

    def _send_full_reply(self, msg, response_text):
        """Deliver replies cleanly, splitting across sequential messages if exceeding Zulip's single-message limit."""
        if not response_text:
            self._send_reply(msg, "*(agent returned no output)*")
            return

        max_chunk = 8500
        if len(response_text) <= max_chunk:
            self._send_reply(msg, response_text)
            return

        # Split into readable chunks (up to 4 parts)
        chunks = []
        remaining = response_text
        while remaining and len(chunks) < 4:
            if len(remaining) <= max_chunk:
                chunks.append(remaining)
                break
            # Find clean split point (paragraph or newline)
            split_idx = remaining.rfind('\n\n', 0, max_chunk)
            if split_idx == -1:
                split_idx = remaining.rfind('\n', 0, max_chunk)
            if split_idx == -1 or split_idx < 2000:
                split_idx = max_chunk
            chunks.append(remaining[:split_idx].rstrip())
            remaining = remaining[split_idx:].lstrip()

        if remaining:
            chunks[-1] += f"\n\n*…[output truncated at {len(response_text)} chars]*"

        for i, chunk in enumerate(chunks):
            header = f"*(Part {i+1}/{len(chunks)})*\n\n" if len(chunks) > 1 else ""
            self._send_reply(msg, header + chunk)
            if i < len(chunks) - 1:
                time.sleep(0.5)

    def _is_long_job(self, content):
        """Detect if a request is a long-running, deep research, or complex task."""
        c = content.strip().lower()
        if c.startswith(('/long', '/deep', '/job', '--long', '--deep', ':long')):
            return True
        long_keywords = [
            "deep research", "full refactor", "run benchmark", "long task",
            "take your time", "train model", "design binder", "boltzgen run",
            "investigate thoroughly", "extensive audit", "long job"
        ]
        return any(kw in c for kw in long_keywords)

    def _strip_long_prefix(self, content):
        """Strip command prefixes like /long, /deep from user prompt."""
        c = content.strip()
        for prefix in ('/long', '/deep', '/job', '--long', '--deep', ':long'):
            if c.lower().startswith(prefix):
                return c[len(prefix):].strip()
        return c

    def _process_message(self, msg, content):
        """Run the ai agent and send the result back. Runs in a background thread."""
        tid = threading.get_ident()
        print(f"[thread-{tid}] Processing: {content[:80]}")

        is_long = self._is_long_job(content)
        clean_content = self._strip_long_prefix(content)

        # Timeouts: default 1800s (30 mins), long jobs 7200s (2 hours)
        default_timeout = int(os.environ.get("BRIDGE_TASK_TIMEOUT", os.environ.get("INFER_TASK_TIMEOUT", 1800)))
        long_timeout = int(os.environ.get("BRIDGE_LONG_TASK_TIMEOUT", os.environ.get("INFER_LONG_TASK_TIMEOUT", 7200)))
        task_timeout = long_timeout if is_long else default_timeout

        # If it's a long job, send an immediate confirmation on Zulip and add a reaction
        if is_long:
            try:
                if hasattr(self.client, "add_reaction"):
                    self.client.add_reaction({
                        "message_id": msg.get("id"),
                        "emoji_name": "hourglass_flowing_sand"
                    })
            except Exception:
                pass
            
            self._send_reply(
                msg,
                f"⏳ **Task confirmed in extended execution mode** (Timeout: {task_timeout}s | Unbounded Steps).\n"
                f"I am executing this task now and will post the full results here when completed."
            )

        # Fetch context messages and manage context window
        context_messages = self._get_context_messages(msg)
        context_messages = self._manage_context_window(context_messages, clean_content)
        prompt = self._construct_prompt_with_context(msg, clean_content, context_messages)

        # Build environment for the subprocess, incorporating ~/.local/share/ai/env
        ai_mode = self._ai_mode()
        run_env = os.environ.copy()
        for k, v in load_env_file().items():
            if k not in run_env or not run_env[k]:
                run_env[k] = v

        if ai_mode == "auto":
            run_env["INFER_AUTO_APPROVE"] = "1"
        else:
            run_env.pop("INFER_AUTO_APPROVE", None)

        sender_email = msg.get("sender_email")
        if sender_email:
            run_env["AI_REMINDER_ZULIP_TO"] = sender_email

        mode_flags = []
        if ai_mode == "auto":
            mode_flags = ["--auto"]
        elif ai_mode == "manual":
            mode_flags = ["--manual"]
        else:
            mode_flags = ["--plan"]

        if is_long:
            mode_flags.append("-c")  # continue until task_complete without bounding at 30/60 steps

        ai_bin = self._resolve_ai_bin()
        ai_cmd = [ai_bin, "-q"] + mode_flags + [prompt]

        try:
            result = subprocess.run(
                ai_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                env=run_env,
                text=True,
                timeout=task_timeout
            )
            response_text = result.stdout
            if result.returncode != 0 and result.stderr:
                response_text += f"\n\n*Stderr:*\n```\n{result.stderr}\n```"

            response_text = clean_response(response_text)
            if not response_text:
                response_text = "*(agent returned no output)*"

        except subprocess.TimeoutExpired:
            response_text = (
                f"⏱️ The agent timed out after {task_timeout}s. "
                "For very long-running jobs, use `/long <prompt>` to run in extended execution mode "
                "or ask me to schedule a background task."
            )
        except Exception as e:
            response_text = f"⚠️ Failed to run local `ai` CLI: {str(e)}"

        self._send_full_reply(msg, response_text)
        print(f"[thread-{tid}] Done.")

    def handle_message(self, msg):
        """Handle a Zulip message."""
        sender_email = msg['sender_email']

        # Don't respond to our own messages
        if sender_email == self.bot_email:
            return

        # Restrict to the owner/configured user to protect privacy
        import getpass
        allowed_user = os.environ.get("ZULIP_USER") or self.detected_owner or getpass.getuser()

        # Normalize values for case-insensitive comparison
        allowed_user_clean = allowed_user.strip().lower()
        sender_email_clean = sender_email.strip().lower()
        sender_full_name_clean = msg.get('sender_full_name', '').strip().lower()
        sender_username_clean = sender_email_clean.split('@')[0] if '@' in sender_email_clean else sender_email_clean

        is_allowed = (
            allowed_user_clean == sender_email_clean or
            allowed_user_clean == sender_full_name_clean or
            allowed_user_clean == sender_username_clean or
            (allowed_user_clean.isdigit() and sender_username_clean == f"user{allowed_user_clean}") or
            (sender_username_clean.startswith("user") and sender_username_clean[4:] == allowed_user_clean)
        )

        print(f"Privacy validation - allowed: '{allowed_user_clean}', sender: '{sender_email_clean}', name: '{sender_full_name_clean}', user: '{sender_username_clean}' -> ALLOWED={is_allowed}")

        if not is_allowed:
            # Quietly ignore messages from others to protect privacy
            return

        content = msg['content'].strip()

        # Built-in commands (cheap, no agent round-trip)
        if content.startswith('/ping'):
            self._send_reply(msg, "🟢 Zulip AI Bridge is alive.")
            return
        if content.startswith('/mode'):
            parts = content.split()
            if len(parts) > 1:
                new_mode = parts[1].strip().lower()
                if new_mode in ("auto", "plan", "manual"):
                    os.environ["BRIDGE_AI_MODE"] = new_mode
                    self._send_reply(msg, f"✅ Bridge AI mode set to: **{new_mode}**")
                    return
            self._send_reply(
                msg,
                f"AI mode: **{self._ai_mode()}** "
                "(use `/mode auto`, `/mode plan`, or `/mode manual` to switch, or `/long <task>` for extended jobs)",
            )
            return
        if content.startswith('/timeout'):
            parts = content.split()
            if len(parts) > 1 and parts[1].isdigit():
                os.environ["BRIDGE_TASK_TIMEOUT"] = parts[1]
                self._send_reply(msg, f"✅ Default task timeout set to: **{parts[1]}s**")
                return
            default_to = os.environ.get("BRIDGE_TASK_TIMEOUT", os.environ.get("INFER_TASK_TIMEOUT", "1800"))
            long_to = os.environ.get("BRIDGE_LONG_TASK_TIMEOUT", os.environ.get("INFER_LONG_TASK_TIMEOUT", "7200"))
            self._send_reply(msg, f"⏱️ Timeouts: Standard: **{default_to}s**, Long Jobs: **{long_to}s** (use `/timeout <seconds>` to update)")
            return

        # If the bot is mentioned in a stream, strip the mention syntax (e.g. @**AI Bot**)
        if msg['type'] != 'private' and content.startswith('@**'):
            mention_end = content.find('**')
            if mention_end != -1:
                mention_end_close = content.find('**', mention_end + 2)
                if mention_end_close != -1:
                    content = content[mention_end_close + 2:].strip()

        # Replace Zulip upload URLs with extracted file content
        content, processed = self._file_parser.process_message_urls(content)

        print(f"Received query from {sender_email}: {content}")

        # Dispatch to a daemon thread — the Zulip event loop returns immediately
        # and is ready to receive the next message while this one is being processed.
        t = threading.Thread(
            target=self._process_message,
            args=(msg, content),
            daemon=True
        )
        t.start()

    def _manage_context_window(self, context_messages, current_content):
        """Apply context window management to prevent overflow."""
        # Exclude the bot's own messages from context — the agent already
        # has its own conversation history; including Zulip-sent bot messages
        # would duplicate and confuse context.
        filtered = [
            m for m in context_messages
            if m.get("sender_email", "") != self.bot_email
        ]
        if len(filtered) != len(context_messages):
            logger.info(
                f"Filtered out {len(context_messages) - len(filtered)} bot messages "
                f"from context ({len(context_messages)} → {len(filtered)})"
            )

        window_manager = ContextWindowManager()
        truncated_messages, truncated = window_manager.truncate_context(
            filtered, current_content
        )
        if truncated:
            logger.info(f"Context truncated from {len(filtered)} to {len(truncated_messages)} messages")
        return truncated_messages

    def run(self):
        """Start the Zulip AI Bridge with automatic reconnection."""
        print("🚀 Starting Zulip AI Bridge listener (threaded — concurrent messages supported)...")
        
        max_backoff = 60  # Maximum backoff in seconds
        current_backoff = 1
        
        while True:
            try:
                self.client.call_on_each_message(self.handle_message)
            except (ConnectionError, requests.exceptions.RequestException) as e:
                print(f"⚠️ Connection error: {e}")
                print(f"⏳ Reconnecting in {current_backoff}s...")
                time.sleep(current_backoff)
                current_backoff = min(current_backoff * 2, max_backoff)
            except Exception as e:
                print(f"❌ Unexpected error: {e}")
                print(f"⏳ Restarting in {current_backoff}s...")
                time.sleep(current_backoff)
                current_backoff = min(current_backoff * 2, max_backoff)


class ContextWindowManager:
    """Manages conversation context to stay within AI model's context window."""
    
    def __init__(self, max_tokens=4096, max_messages=10):
        """
        Initialize the context window manager.
        
        Args:
            max_tokens: Maximum estimated tokens for the prompt (conservative estimate)
            max_messages: Maximum number of context messages to include
        """
        self.max_tokens = max_tokens
        self.max_messages = max_messages
    
    def estimate_tokens(self, text):
        """Rough token estimation: ~4 chars per token for English text."""
        return len(text) // 4
    
    def truncate_context(self, context_messages, current_content):
        """
        Truncate context messages to fit within the token / message budget.

        Strategy: keep the most recent messages and drop the oldest ones when
        either the ``max_tokens`` budget or the ``max_messages`` cap is
        exceeded.  This preserves the most relevant conversational context.

        Args:
            context_messages: List of message dictionaries
            current_content: The current message content

        Returns:
            Tuple of (truncated messages, whether truncation occurred)
        """
        if not context_messages:
            return context_messages, False

        # Estimate how many characters the current query occupies.
        estimated_length = len(current_content)
        kept = []

        # Walk backwards through the context so the most recent messages are kept.
        for msg in reversed(context_messages):
            sender = msg.get("sender_full_name", msg.get("sender_email", "Unknown"))
            body = msg.get("content", "").strip()
            msg_length = len(f"- {sender}: {body}")

            # Check token budget (rough estimate: 4 chars ≈ 1 token)
            if estimated_length + msg_length > self.max_tokens * 4:
                logger.info(
                    f"Context truncated to fit token budget: keeping "
                    f"{len(kept)} of {len(context_messages)} messages."
                )
                break

            # Check message count cap
            if len(kept) >= self.max_messages:
                logger.info(
                    f"Context truncated to fit message cap: keeping "
                    f"{len(kept)} of {len(context_messages)} messages."
                )
                break

            kept.append(msg)
            estimated_length += msg_length

        # Restore to chronological order so the prompt reads naturally.
        kept = list(reversed(kept))
        truncated = len(kept) < len(context_messages)
        return kept, truncated
    
    def format_context(self, context_messages, bot_email):
        """Format context messages into a readable string for the prompt."""
        if not context_messages:
            return ""
        
        context_lines = ["---", "Recent conversation context (for reference):"]
        for m in context_messages:
            sender = m.get("sender_full_name", m.get("sender_email"))
            if m.get("sender_email") == bot_email:
                sender = "AI (You)"
            else:
                sender = f"User ({sender})"
            body = m.get("content", "").strip()
            if "\n" in body:
                body = "\n".join("  " + line for line in body.splitlines())
            context_lines.append(f"- {sender}: {body}")
        context_lines.append("---")
        context_lines.append("Latest query/message:")
        
        return "\n".join(context_lines)

if __name__ == "__main__":
    try:
        bridge = ZulipAiBridge()
        bridge.run()
    except Exception as e:
        print(f"Error starting bridge: {e}")
        sys.exit(1)
